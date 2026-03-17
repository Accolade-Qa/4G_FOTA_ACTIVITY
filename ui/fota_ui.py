import json
import os
import sys
from pathlib import Path

from PyQt6.QtCore import QProcess, Qt, QTimer
from PyQt6.QtGui import QIcon
from PyQt6.QtWidgets import (
    QApplication,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QPlainTextEdit,
    QFileDialog,
    QSpinBox,
    QVBoxLayout,
    QWidget,
)


APP_TITLE = "FOTA Automation UI"
ASSETS_DIR = Path(__file__).resolve().parent / "assets"
LOGO_PATH = ASSETS_DIR / "logo.png"
LAST_REPO_PATH_FILE = Path.home() / ".fota_ui_repo_root"


def read_properties(path: Path) -> dict:
    data = {}
    if not path.exists():
        return data
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        data[key.strip()] = value.strip()
    return data


def write_properties(path: Path, data: dict) -> None:
    lines = [
        "# Serial configuration",
        f"serial.port={data.get('serial.port', '')}",
        f"serial.baud={data.get('serial.baud', '115200')}",
        "",
        "# Input/output paths",
        f"firmware.csv={data.get('firmware.csv', 'input/fota_batch.csv')}",
        f"audit.csv={data.get('audit.csv', 'results/fota_audit.csv')}",
        f"firmware.json={data.get('firmware.json', 'input/servers.json')}",
        f"login.json={data.get('login.json', 'results/login_packets.json')}",
        "",
        "# Portal credentials and URL",
        f"login.url={data.get('login.url', '')}",
        f"login.user={data.get('login.user', '')}",
        f"login.pass={data.get('login.pass', '')}",
        "",
        "# Device state (used as default if device reports 'Default')",
        f"state={data.get('state', 'Default')}",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def find_jar(target_dir: Path) -> Path | None:
    if not target_dir.exists():
        return None
    jars = list(target_dir.glob("*.jar"))
    if not jars:
        return None
    preferred = [j for j in jars if "shaded" in j.name and "original" not in j.name]
    if preferred:
        return preferred[0]
    filtered = [j for j in jars if "original" not in j.name]
    return filtered[0] if filtered else None


def _normalize_header(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum())


def generate_servers_json_from_inputs(
    xlsx_path: Path,
    csv_path: Path,
    json_path: Path,
    default_state: str | None = None,
) -> tuple[bool, str]:
    by_state: dict[str, dict[str, object]] = {}
    warnings: list[str] = []

    if xlsx_path.exists():
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            warnings.append(f"Missing dependency 'openpyxl': {exc}")
        else:
            wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
            if wb.sheetnames:
                for sheet_name in wb.sheetnames:
                    if sheet_name.strip().lower() == "summary":
                        continue
                    ws = wb[sheet_name]
                    rows = ws.iter_rows(min_row=1, values_only=True)
                    try:
                        header_row = next(rows)
                    except StopIteration:
                        continue

                    headers = []
                    for cell in header_row:
                        headers.append("" if cell is None else _normalize_header(str(cell)))

                    fw_idx = None
                    for i, name in enumerate(headers):
                        if not name:
                            continue
                        if "firmwarename" in name or name in ("firmware", "version"):
                            fw_idx = i
                            break
                    if fw_idx is None:
                        fw_idx = 1  # default to second column

                    versions: list[str] = []
                    for row in rows:
                        if row is None:
                            continue
                        if fw_idx >= len(row):
                            continue
                        raw = row[fw_idx]
                        if raw is None:
                            continue
                        version = str(raw).strip()
                        if not version:
                            continue
                        versions.append(version)

                    if versions:
                        entry = by_state.setdefault(
                            sheet_name.strip(),
                            {"state": sheet_name.strip(), "versions": []},
                        )
                        entry_versions = entry.setdefault("versions", [])
                        for v in versions:
                            if v not in entry_versions:
                                entry_versions.append(v)
            else:
                warnings.append("No sheets found in Excel file.")
    else:
        warnings.append(f"Server input not found: {xlsx_path}")

    if csv_path.exists():
        import csv

        with csv_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                warnings.append("Firmware CSV has no headers.")
            else:
                headers = {_normalize_header(h): h for h in reader.fieldnames if h is not None}
                state_header = headers.get("state")
                version_header = (
                    headers.get("ufw")
                    or headers.get("version")
                    or headers.get("firmwareversion")
                    or headers.get("swversion")
                )

                if not version_header:
                    warnings.append("Firmware CSV missing required column: UFW/version.")
                else:
                    for row in reader:
                        if not row:
                            continue
                        version_raw = row.get(version_header)
                        state_raw = row.get(state_header) if state_header else None

                        version = str(version_raw).strip() if version_raw is not None else ""
                        state = str(state_raw).strip() if state_raw is not None else ""
                        if not state:
                            state = default_state or ""
                        if not state or not version:
                            continue

                        entry = by_state.setdefault(state, {"state": state, "versions": []})
                        versions = entry.setdefault("versions", [])
                        if version not in versions:
                            versions.append(version)
    else:
        warnings.append(f"Firmware CSV not found: {csv_path}")

    if not by_state:
        details = "; ".join(warnings) if warnings else "No input data."
        return False, f"No server data generated. {details}"

    data = list(by_state.values())
    for item in data:
        if not item.get("versions"):
            item["versions"] = []

    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(data, indent=2), encoding="utf-8")

    detail = "; ".join(warnings) if warnings else "OK"
    return True, f"Generated {json_path} from inputs. {detail}"




class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        if LOGO_PATH.exists():
            self.setWindowIcon(QIcon(str(LOGO_PATH)))
        self.resize(980, 720)

        self.repo_root = self._resolve_repo_root()
        self._apply_repo_root(self.repo_root)

        self.backend_process: QProcess | None = None
        self.build_process: QProcess | None = None
        self.stop_requested = False

        self._init_ui()
        self._load_config()
        QTimer.singleShot(0, self._start_backend)

    def _apply_repo_root(self, repo_root: Path) -> None:
        self.repo_root = repo_root
        self.config_path = repo_root / "config.properties"
        self.target_dir = repo_root / "target"
        self.input_dir = repo_root / "input"
        self.server_inputs_xlsx = self.input_dir / "Server_Inputs.xlsx"

    def _is_repo_root(self, path: Path) -> bool:
        if not path.exists() or not path.is_dir():
            return False
        if (path / "pom.xml").exists():
            return True
        if (path / "config.properties").exists() and (path / "src").exists():
            return True
        return False

    def _resolve_repo_root(self) -> Path:
        candidates: list[Path] = []
        if LAST_REPO_PATH_FILE.exists():
            try:
                saved = Path(LAST_REPO_PATH_FILE.read_text(encoding="utf-8").strip())
                if saved:
                    candidates.append(saved)
            except Exception:
                pass
        candidates.append(Path(__file__).resolve().parents[1])
        candidates.append(Path.cwd())

        for candidate in candidates:
            if self._is_repo_root(candidate):
                return candidate

        chosen = QFileDialog.getExistingDirectory(None, "Select FOTA repo folder", str(Path.cwd()))
        if chosen:
            chosen_path = Path(chosen)
            if self._is_repo_root(chosen_path):
                try:
                    LAST_REPO_PATH_FILE.write_text(str(chosen_path), encoding="utf-8")
                except Exception:
                    pass
                return chosen_path

        return Path(__file__).resolve().parents[1]

    def _init_ui(self) -> None:
        root = QWidget()
        layout = QVBoxLayout(root)

        form = QFormLayout()

        self.serial_port = QLineEdit()
        self.serial_port.setPlaceholderText("COM3 or leave empty for auto-detect")
        form.addRow("Serial Port", self.serial_port)

        self.baud_rate = QSpinBox()
        self.baud_rate.setRange(1200, 1000000)
        self.baud_rate.setValue(115200)
        form.addRow("Baud Rate", self.baud_rate)

        self.firmware_json = QLineEdit()
        form.addRow("Servers JSON", self._with_browse(self.firmware_json, "Select servers.json"))

        self.firmware_csv = QLineEdit()
        form.addRow("Firmware CSV", self._with_browse(self.firmware_csv, "Select firmware CSV"))

        self.audit_csv = QLineEdit()
        form.addRow("Audit CSV", self._with_browse(self.audit_csv, "Select audit CSV"))

        self.login_json = QLineEdit()
        form.addRow("Login JSON", self._with_browse(self.login_json, "Select login JSON"))

        self.portal_url = QLineEdit()
        form.addRow("Portal URL", self.portal_url)

        self.portal_user = QLineEdit()
        form.addRow("Portal User", self.portal_user)

        self.portal_pass = QLineEdit()
        self.portal_pass.setEchoMode(QLineEdit.EchoMode.Password)
        form.addRow("Portal Pass", self.portal_pass)

        self.default_state = QLineEdit()
        form.addRow("Default State", self.default_state)

        layout.addLayout(form)

        button_row = QHBoxLayout()
        self.save_btn = QPushButton("Save Config")
        self.save_btn.clicked.connect(self._save_config)
        self.start_btn = QPushButton("Start Backend")
        self.start_btn.clicked.connect(self._start_backend)
        self.stop_btn = QPushButton("Stop Backend")
        self.stop_btn.clicked.connect(self._stop_backend)
        self.stop_btn.setEnabled(False)

        button_row.addWidget(self.save_btn)
        button_row.addWidget(self.start_btn)
        button_row.addWidget(self.stop_btn)
        button_row.addStretch(1)
        layout.addLayout(button_row)

        layout.addWidget(QLabel("Backend Log"))
        self.log = QPlainTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log, stretch=1)

        self.setCentralWidget(root)

    def _with_browse(self, line_edit: QLineEdit, title: str) -> QWidget:
        wrapper = QWidget()
        row = QHBoxLayout(wrapper)
        row.setContentsMargins(0, 0, 0, 0)
        row.addWidget(line_edit, stretch=1)
        browse = QPushButton("Browse")
        browse.clicked.connect(lambda: self._browse_file(line_edit, title))
        row.addWidget(browse)
        return wrapper

    def _browse_file(self, line_edit: QLineEdit, title: str) -> None:
        path, _ = QFileDialog.getOpenFileName(self, title, str(self.repo_root))
        if path:
            rel = os.path.relpath(path, self.repo_root)
            line_edit.setText(rel)

    def _load_config(self) -> None:
        props = read_properties(self.config_path)
        def _get(key: str, default: str) -> str:
            value = props.get(key, "").strip()
            return value if value else default

        self.serial_port.setText(props.get("serial.port", ""))
        self.baud_rate.setValue(int(_get("serial.baud", "115200")))
        self.firmware_json.setText(_get("firmware.json", "input/servers.json"))
        self.firmware_csv.setText(_get("firmware.csv", "input/fota_batch.csv"))
        self.audit_csv.setText(_get("audit.csv", "results/fota_audit.csv"))
        self.login_json.setText(_get("login.json", "results/login_packets.json"))
        self.portal_url.setText(props.get("login.url", ""))
        self.portal_user.setText(props.get("login.user", ""))
        self.portal_pass.setText(props.get("login.pass", ""))
        self.default_state.setText(_get("state", "Default"))

    def _collect_config(self) -> dict:
        return {
            "serial.port": self.serial_port.text().strip(),
            "serial.baud": str(self.baud_rate.value()),
            "firmware.csv": self.firmware_csv.text().strip(),
            "audit.csv": self.audit_csv.text().strip(),
            "firmware.json": self.firmware_json.text().strip(),
            "login.json": self.login_json.text().strip(),
            "login.url": self.portal_url.text().strip(),
            "login.user": self.portal_user.text().strip(),
            "login.pass": self.portal_pass.text().strip(),
            "state": self.default_state.text().strip(),
        }

    def _save_config(self) -> None:
        write_properties(self.config_path, self._collect_config())
        self._append_log(f"Saved config to {self.config_path}")

    def _start_backend(self) -> None:
        if self.backend_process and self.backend_process.state() != QProcess.ProcessState.NotRunning:
            QMessageBox.information(self, APP_TITLE, "Backend is already running.")
            return
        if self.build_process and self.build_process.state() != QProcess.ProcessState.NotRunning:
            QMessageBox.information(self, APP_TITLE, "Build is already running.")
            return

        self.stop_requested = False
        self._save_config()
        self._generate_servers_json()

        jar = find_jar(self.target_dir)
        if not jar:
            self._append_log("No JAR found in target/. Running Maven package...")
            self._run_maven_build()
            return

        self._run_java(jar)

    def _run_maven_build(self) -> None:
        if self.build_process and self.build_process.state() != QProcess.ProcessState.NotRunning:
            return
        self.build_process = QProcess(self)
        self.build_process.setWorkingDirectory(str(self.repo_root))
        self.build_process.readyReadStandardOutput.connect(
            lambda: self._append_process_output(self.build_process)
        )
        self.build_process.readyReadStandardError.connect(
            lambda: self._append_process_output(self.build_process)
        )
        self.build_process.finished.connect(self._on_build_finished)
        self.build_process.errorOccurred.connect(
            lambda err: self._on_process_error("Build", err)
        )
        self._append_log("Running Maven build: mvn -q -DskipTests package")
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(False)
        self.build_process.start("mvn", ["-q", "-DskipTests", "package"])

    def _on_build_finished(self, exit_code: int, exit_status: QProcess.ExitStatus) -> None:
        self.build_process = None
        if self.stop_requested:
            self._append_log("Build stopped.")
            self.start_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)
            return
        if exit_status != QProcess.ExitStatus.NormalExit or exit_code != 0:
            self._append_log(f"Build failed (exit code {exit_code}).")
            self.start_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)
            return
        jar = find_jar(self.target_dir)
        if not jar:
            self._append_log("Build finished but no JAR was found in target/.")
            self.start_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)
            return
        self._run_java(jar)

    def _run_java(self, jar: Path) -> None:
        self.backend_process = QProcess(self)
        self.backend_process.setWorkingDirectory(str(self.repo_root))
        self.backend_process.readyReadStandardOutput.connect(
            lambda: self._append_process_output(self.backend_process)
        )
        self.backend_process.readyReadStandardError.connect(
            lambda: self._append_process_output(self.backend_process)
        )
        self.backend_process.finished.connect(self._on_backend_finished)
        self.backend_process.errorOccurred.connect(
            lambda err: self._on_process_error("Backend", err)
        )
        self._append_log(f"Starting backend: {jar.name}")
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.backend_process.start(
            "java",
            ["-Dfota.config=config.properties", "-jar", str(jar)],
        )

    def _generate_servers_json(self) -> None:
        json_path = self._resolve_path(self.firmware_json.text().strip())
        ok, message = generate_servers_json_from_inputs(
            self.server_inputs_xlsx,
            self._resolve_path(self.firmware_csv.text().strip()),
            json_path,
            self.default_state.text().strip(),
        )
        if ok:
            self._append_log(message)
        else:
            self._append_log(f"[WARN] {message}")

    def _resolve_path(self, value: str) -> Path:
        if not value:
            return self.repo_root
        path = Path(value)
        return path if path.is_absolute() else (self.repo_root / path)

    def _stop_backend(self) -> None:
        self.stop_requested = True
        if self.build_process and self.build_process.state() != QProcess.ProcessState.NotRunning:
            self._stop_process(self.build_process, "Build")
            self.build_process = None
        if self.backend_process and self.backend_process.state() != QProcess.ProcessState.NotRunning:
            self._stop_process(self.backend_process, "Backend")
            self.backend_process = None
            self._append_log("Backend stopped.")
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

    def _on_backend_finished(self, exit_code: int, exit_status: QProcess.ExitStatus) -> None:
        self.backend_process = None
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        if self.stop_requested:
            return
        if exit_status == QProcess.ExitStatus.NormalExit:
            self._append_log(f"Backend exited (code {exit_code}).")
        else:
            self._append_log("Backend crashed.")

    def _append_process_output(self, proc: QProcess | None) -> None:
        if proc is None:
            return
        data = proc.readAllStandardOutput().data().decode(errors="ignore")
        err = proc.readAllStandardError().data().decode(errors="ignore")
        if data:
            self._append_log(data.rstrip())
        if err:
            self._append_log(err.rstrip())

    def _on_process_error(self, label: str, err: QProcess.ProcessError) -> None:
        self._append_log(f"{label} process error: {err.name}")
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

    def _append_log(self, text: str) -> None:
        if not text:
            return
        self.log.appendPlainText(text)

    def _stop_process(self, proc: QProcess, label: str) -> None:
        proc.terminate()
        proc.waitForFinished(5000)
        if proc.state() != QProcess.ProcessState.NotRunning:
            self._append_log(f"{label} did not stop on terminate; killing...")
            proc.kill()
            proc.waitForFinished(3000)
        if proc.state() != QProcess.ProcessState.NotRunning:
            pid = proc.processId()
            if pid:
                try:
                    if sys.platform.startswith("win"):
                        QProcess.execute("taskkill", ["/PID", str(pid), "/T", "/F"])
                    else:
                        QProcess.execute("kill", ["-9", str(pid)])
                    proc.waitForFinished(3000)
                except Exception as exc:
                    self._append_log(f"{label} force kill failed: {exc}")

    def closeEvent(self, event) -> None:
        self._stop_backend()
        event.accept()


def main() -> int:
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
